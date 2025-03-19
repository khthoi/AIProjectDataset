import subprocess
import sys
import os
import time
# Hàm kiểm tra và cài đặt thư viện nếu chưa có
def install_packages():
    try:
        import openpyxl
        import google.generativeai as genai
        import dotenv
    except ImportError:
        print("🔧 Đang cài đặt thư viện cần thiết...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "google-generativeai", "python-dotenv"])
        print("✅ Cài đặt thành công! Chạy lại chương trình...")
        sys.exit()

# Kiểm tra và cài đặt thư viện
install_packages()
import openpyxl
from dotenv import load_dotenv
import google.generativeai as genai
# Load API Key từ file .env
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")
if not API_KEY:
    print("❌ Lỗi: Không tìm thấy API Key trong file .env. Vui lòng thêm GEMINI_API_KEY vào .env")
    sys.exit(1)

# Cấu hình API Gemini
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel("gemini-2.0-flash")

# Đọc file Excel
file_path = "chuanhoa.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Hàm gọi API Gemini để chuẩn hóa văn bản
def normalize_text(text, max_retries=5, delay=2):
    prompt = f"""Bạn là một chuyên gia xử lý ngôn ngữ tự nhiên. Hãy chuẩn hóa và viết lại đoạn sau thành tiếng việt theo yêu cầu:
    1. Nếu đoạn văn có nhiều ý câu, gạch đầu dòng hoặc nhiều đoạn nhỏ, hãy gộp lại để chỉ còn 2 - 3 đoạn.
    2. Không được cắt bớt từ ngữ, đảm bảo giữ nguyên toàn bộ nội dung.
    3. Chuẩn hóa chính tả, sửa các lỗi chính tả hoặc ký tự lỗi, lỗi font chữ, lỗi dấu câu.
    4. Đặc biệt đoạn văn sau khi chuẩn hóa phải trên 300 từ, không quá 800 từ.
    5. Chỉ cần sinh ra văn bản, không cần thêm tiêu đề như "Văn bản đã được chuẩn hóa như sau" hay "Dưới đây là văn bản đã được chuẩn hóa".
    Văn bản cần chuẩn hóa:
    {text}
    """
    
    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            normalized_text = response.text.strip()
            if normalized_text:
                return normalized_text
        except Exception as e:
            print(f"Lần thử {attempt + 1}/{max_retries} thất bại: {e}. Thử lại sau {delay} giây...")
            wb.save(file_path)  # Lưu file Excel trước khi dừng
            time.sleep(delay)
    
    return "Không thể chuẩn hóa văn bản sau nhiều lần thử"

# Duyệt từng hàng trong cột 1, gửi API và lưu vào cột 2
for row in range(2, sheet.max_row + 1):  # Bỏ qua hàng tiêu đề (nếu có)
    input_text = sheet.cell(row=row, column=1).value

    if input_text is None or input_text.strip() == "":  # Nếu gặp ô trống, dừng chương trình
        print(f"🚀 Hết dữ liệu cần xử lý tại dòng {row}. Dừng chương trình.")
        break

    print(f"📌 Đang chuẩn hóa dòng {row}...")
    normalized_text = normalize_text(input_text)
    sheet.cell(row=row, column=2, value=normalized_text)
    print(f"✅ Dòng {row} đã chuẩn hóa: {normalized_text}")

# Lưu file Excel
wb.save(file_path)
print("🎉 Hoàn thành! Dữ liệu đã được cập nhật vào file chuanhoa.xlsx")