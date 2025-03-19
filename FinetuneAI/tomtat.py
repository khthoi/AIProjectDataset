import subprocess
import sys
import os
import time
# Hàm kiểm tra và cài đặt thư viện nếu chưa có
def install_packages():
    try:
        import openpyxl
        import google.generativeai as genai
        import dotenv
    except ImportError:
        print("🔧 Đang cài đặt thư viện cần thiết...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "google-generativeai", "python-dotenv"])
        print("✅ Cài đặt thành công! Chạy lại chương trình...")
        sys.exit()

# Kiểm tra và cài đặt thư viện
install_packages()
import openpyxl
from dotenv import load_dotenv
import google.generativeai as genai
# Load API Key từ file .env
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")
if not API_KEY:
    print("❌ Lỗi: Không tìm thấy API Key trong file .env. Vui lòng thêm GEMINI_API_KEY vào .env")
    sys.exit(1)

# Cấu hình API Gemini
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel("gemini-2.0-flash")

# Đọc file Excel
file_path = "tomtat.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Hàm gọi API Gemini để chuẩn hóa văn bản
def normalize_text(text, max_retries=5, delay=2):
    prompt = f"""Bạn là một hệ thống tóm tắt văn bản. Nhiệm vụ của bạn là tóm tắt nội dung một cách chính xác và hiệu quả. Dưới đây là hướng dẫn từng bước để bạn thực hiện nhiệm vụ này:
 1. Đọc kỹ văn bản được cung cấp để nắm bắt nội dung tổng thể. Chú ý đến các câu chủ đề và phần văn bản nhấn mạnh để xác định các ý chính cần thiết.
 2. Phân loại thông tin trong văn bản theo mức độ quan trọng. Ưu tiên những ý chính có tác động lớn đến nội dung tổng thể và loại bỏ các chi tiết phụ hoặc không cần thiết.
 3. Sử dụng các ý chính đã xác định để viết một đoạn tóm tắt ngắn, có độ dài từ 2 đến 3 câu.
 4. Đảm bảo rằng đoạn tóm tắt có các thuộc tính sau:
    - Tính trung thực: Đảm bảo tất cả nội dung của đoạn tóm tắt đều xuất phát từ văn bản gốc mà không thêm thông tin mới.
    - Tính mạch lạc: Trình bày nội dung một cách rõ ràng và nhất quán, đảm bảo các câu trong đoạn tóm tắt kết nối logic với nhau.
    - Tính liên quan: Chỉ bao gồm những ý quan trọng từ văn bản gốc, loại bỏ chi tiết không cần thiết hoặc không liên quan.
 5. Kiểm tra lại đoạn tóm tắt để đảm bảo nó đáp ứng các tiêu chí trên. Thực hiện sửa đổi nếu cần thiết để cải thiện chất lượng tóm tắt.
 6. Văn bản được tóm tắt phải nằm trong khoảng từ 95 cho đến 130 từ.
 Làm theo các bước trên để đưa ra một đoạn tóm tắt chính xác và hiệu quả.
Văn bản tóm tắt là:
    {text}
    """
    
    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            normalized_text = response.text.strip()
            if normalized_text:
                return normalized_text
        except Exception as e:
            print(f"Lần thử {attempt + 1}/{max_retries} thất bại: {e}. Thử lại sau {delay} giây...")
            wb.save(file_path)  # Lưu file Excel trước khi dừng
            time.sleep(delay)
    
    return "Không thể tóm tắt văn bản sau nhiều lần thử"

# Duyệt từng hàng trong cột 2, gửi API và lưu vào cột 3
for row in range(2, sheet.max_row + 1):  # Bỏ qua hàng tiêu đề (nếu có)
    input_text = sheet.cell(row=row, column=2).value  # Đọc từ cột thứ 2

    if input_text is None or input_text.strip() == "":  # Nếu gặp ô trống, dừng chương trình
        print(f"🚀 Hết dữ liệu cần xử lý tại dòng {row}. Dừng chương trình.")
        break

    print(f"📌 Đang tóm tắt dòng {row}...")
    normalized_text = normalize_text(input_text)
    sheet.cell(row=row, column=3, value=normalized_text)  # Ghi vào cột thứ 3
    print(f"✅ Dòng {row} đã tóm tắt: {normalized_text}")
    time.sleep(3)  # Đợi 4 giây trước khi gửi yêu cầu tiếp theo, tránh quá tải API

# Lưu file Excel
wb.save(file_path)
print("🎉 Hoàn thành! Dữ liệu đã được cập nhật vào file tomtat.xlsx")
